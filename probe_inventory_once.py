#!/usr/bin/env python3
"""Round-2 probe: cover every inventory row NOT already covered by
probe_sources_once.py (built-in adapters), then annotate the source
inventory xlsx with local-network availability (direct vs proxy).
"""

from __future__ import annotations

import json
import sys
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from pathlib import Path

import feedparser
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

PROXY = "http://127.0.0.1:7897"
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")


def make_session(proxy: str | None) -> requests.Session:
    s = requests.Session()
    s.trust_env = False
    if proxy:
        s.proxies = {"http": proxy, "https": proxy}
    return s


# (label, url, kind) kind: feed | page | api
TARGETS = [
    # --- OPML example feeds (new ones; the 4 overlaps reuse round-1) ---
    ("Google AI Blog (OPML URL)", "https://blog.google/technology/ai/rss/", "feed"),
    ("Microsoft AI Blog", "https://news.microsoft.com/source/topics/ai/feed/", "feed"),
    ("Wired AI", "https://www.wired.com/feed/tag/ai/latest/rss", "feed"),
    ("InfoQ CN", "https://www.infoq.cn/feed", "feed"),
    ("NVIDIA Generative AI Blog", "https://developer.nvidia.com/blog/category/generative-ai/feed/", "feed"),
    ("Baoyu", "https://baoyu.io/feed.xml", "feed"),
    ("Simon Willison", "https://simonwillison.net/atom/everything/", "feed"),
    # --- RSSHub replacement targets (new ones) ---
    ("Readhub", "https://readhub.cn/rss", "feed"),
    ("36kr", "https://36kr.com/feed", "feed"),
    ("Sspai", "https://sspai.com/feed", "feed"),
    ("Meituan Tech", "https://tech.meituan.com/feed", "feed"),
    ("mjg59 dreamwidth", "http://mjg59.dreamwidth.org/data/rss", "feed"),
    # --- bridges ---
    ("Telegram bridge (sample: durov)", "https://t.me/s/durov", "page"),
    ("Jike bridge (host)", "https://m.okjike.com/", "page"),
    # --- candidate rsshub instance ---
    ("Karpathy X (rsshub.pseudoyu.com)", "https://rsshub.pseudoyu.com/twitter/user/karpathy", "feed"),
    # --- paid API hosts (expect auth errors without keys = reachable) ---
    ("X API", "https://api.x.com/2/tweets/search/recent?query=AI&max_results=10", "api"),
    ("SocialData.tools", "https://api.socialdata.tools/search/recent/ai", "api"),
    ("TikHub", "https://api.tikhub.io/", "api"),
    ("AgentMail", "https://api.agentmail.to/", "api"),
    # --- auxiliary services ---
    ("Jina Reader", "https://r.jina.ai/https://aibreakfast.beehiiv.com/", "page"),
    ("Google Translate", "https://translate.googleapis.com/translate_a/single?client=gtx&sl=en&tl=zh-CN&dt=t&q=hello", "api"),
    ("DeepSeek API", "https://api.deepseek.com/chat/completions", "api"),
    # --- evaluated, not adopted ---
    ("OpenRouter Announcements", "https://openrouter.ai/blog/feed.xml", "feed"),
    ("LMSYS Blog", "https://lmsys.org/feed.xml", "feed"),
    ("HF Daily Papers (JSON API)", "https://huggingface.co/api/daily_papers", "api"),
    ("Berkeley RDI Blog (guessed)", "https://rdi.berkeley.edu/feed.xml", "feed"),
    ("QbitAI", "https://www.qbitai.com/feed", "feed"),
    ("Substack (latent.space, sample)", "https://www.latent.space/feed", "feed"),
]


def probe_one(sess: requests.Session, label: str, url: str, kind: str) -> dict:
    t0 = time.perf_counter()
    row = {"label": label, "url": url, "kind": kind, "ok": False,
           "status": None, "entries": None, "bytes": 0, "ms": 0, "error": ""}
    try:
        r = sess.get(url, timeout=15, headers={"User-Agent": UA}, allow_redirects=True)
        row["status"] = r.status_code
        row["bytes"] = len(r.content)
        if kind == "feed" and r.status_code == 200:
            parsed = feedparser.parse(r.content)
            row["entries"] = len(parsed.entries)
            row["ok"] = row["entries"] > 0
        elif kind in ("page", "api"):
            row["ok"] = r.status_code < 500
    except Exception as exc:
        msg = str(exc).splitlines()[0][:100]
        row["error"] = f"{type(exc).__name__}: {msg}"
    row["ms"] = int((time.perf_counter() - t0) * 1000)
    return row


def probe_mode(mode: str, proxy: str | None) -> list[dict]:
    sess = make_session(proxy)
    rows = []
    with ThreadPoolExecutor(max_workers=6) as pool:
        futures = [pool.submit(probe_one, sess, l, u, k) for l, u, k in TARGETS]
        for f in futures:
            r = f.result()
            r["mode"] = mode
            rows.append(r)
            if r["kind"] == "feed":
                detail = f"{r['entries']} entries" if r["entries"] is not None else (r["error"] or f"HTTP {r['status']}")
            else:
                detail = f"HTTP {r['status']}, {r['bytes']}B" if r["status"] else r["error"]
            print(f"[{mode}] {'OK  ' if r['ok'] else 'FAIL'} {r['label']:<36} {r['ms']:>6}ms  {detail}", flush=True)
    return rows


def cell_text(row: dict) -> str:
    if row is None:
        return "—"
    if row["kind"] == "feed":
        if row["ok"]:
            return f"✓ {row['entries']} entries"
        if row["status"]:
            return f"✗ HTTP {row['status']}"
        return f"✗ {row['error'].split(':')[0]}"
    if row["ok"]:
        return f"✓ HTTP {row['status']}"
    if row["status"]:
        return f"✗ HTTP {row['status']}"
    return f"✗ {row['error'].split(':')[0]}"


def main() -> int:
    round2 = probe_mode("direct", None) + probe_mode("proxy", PROXY)
    (ROOT / "probe-inventory-report.json").write_text(
        json.dumps(round2, ensure_ascii=False, indent=1), encoding="utf-8")

    by_label = {}
    for r in round2:
        by_label[(r["mode"], r["label"])] = r

    # ---- round-1 results (built-in adapters) summarized from probe-sources-report.json ----
    r1: dict = {}
    p1 = ROOT / "probe-sources-report.json"
    if p1.exists():
        for r in json.loads(p1.read_text(encoding="utf-8")):
            r1[(r["mode"], r["name"])] = r

    def r1_text(mode: str, name: str) -> str:
        row = r1.get((mode, name))
        if not row:
            return "—"
        if row["ok"]:
            return f"✓ {row['items']} items"
        return "✗ timeout" if "Timeout" in row["error"] else f"✗ {row['error'][:28]}"

    # (category, sheet-name) -> (direct_text, proxy_text, note)
    def feed_pair(label: str):
        return cell_text(by_label.get(("direct", label))), cell_text(by_label.get(("proxy", label)))

    R1_MAP = {
        ("OpenAI News", "OpenAI News"),
        ("Google DeepMind", "Google DeepMind"),
        ("Hugging Face Blog", "Hugging Face Blog"),
        ("GitHub AI & ML", "GitHub AI & ML"),
        ("OpenAI Skills", "OpenAI Skills"),
        ("The Decoder AI News", "The Decoder AI News"),
        ("TechCrunch AI", "TechCrunch AI"),
        ("The Verge", "The Verge"),
        ("MarkTechPost Research", "MarkTechPost Research"),
        ("VentureBeat AI", "VentureBeat AI"),
        ("Artificial Intelligence News", "Artificial Intelligence News"),
        ("Claude Code Releases", "Claude Code Releases"),
        ("Anthropic News", "Anthropic News (page)"),
        ("OpenAI Codex Changelog", "OpenAI Codex Changelog (page)"),
        ("AI HOT", "AI HOT"),
        ("AI Breakfast", "AI Breakfast"),
        ("Follow Builders", "Follow Builders"),
        ("AI HubToday", "AI HubToday"),
        ("AIbase", "AIbase"),
        ("TechURLs", "TechURLs"),
        ("Buzzing", "Buzzing"),
        ("Info Flow (Iris)", "Info Flow (iris)"),
        ("BestBlogs", "BestBlogs"),
        ("Zeli（HN 24h 最热）", "Zeli"),
        ("Hacker News Algolia", "Hacker News (Algolia)"),
        ("NewsNow", "NewsNow"),
        ("WaytoAGI", "WayToAGI (feishu)"),
    }

    NOTES = {
        "OpenAI Codex Changelog": "连通正常；45 天窗口内暂无新条目",
        "Anthropic News": "页面解析正常",
        "Info Flow (Iris)": "可用但很慢（约 46s）",
        "WaytoAGI": "可用，近 7 日 35 条更新",
        "AI Breakfast": "源站问题：r.jina.ai 全面 403，云端管线同样失败，非本地网络问题",
        "NewsNow": "本地 IP 与代理出口均被 403 拦截；云端 Actions 正常，仅影响本地运行",
        "Hugging Face Blog": "github.com/huggingface.co 直连不通，需代理",
        "OpenAI Skills": "走代理连通；0 条是关键词过滤未命中，属正常",
        "Claude Code Releases": "github.com 直连不通，需代理",
        "OPML 私有订阅（入口）": "本地文件 feeds/follow.opml（未入库），非网络探测对象",
        "RSSHub 路由替换规则": "规则条目；替换目标已逐个探测（InfoQ/HF/Readhub/36kr/少数派/美团技术）",
        "OPML 跳过路由": "设计上跳过（桥接不稳定或无官方 RSS），不探测",
        "Telegram 桥接": "以示例频道 @durov 探测；实际频道取决于私有 OPML",
        "即刻桥接": "探测 m.okjike.com 主站可达性；具体话题页取决于私有 OPML",
        "X API 官方近期搜索": "网络可达性探测（无密钥，401 属正常）；启用需 X_BEARER_TOKEN",
        "SocialData.tools X 搜索 + 列表": "网络可达性探测（无密钥）；启用需 SOCIALDATA_API_KEY",
        "TikHub 抖音 / 小红书搜索": "网络可达性探测（无密钥）；启用需 TIKHUB_API_KEY",
        "AgentMail 邮件摘要": "网络可达性探测（无密钥）；启用需 AGENTMAIL_API_KEY",
        "Jina Reader": "AI Breakfast 与正文兜底依赖；当前对本地返回 403",
        "Google Translate": "翻译链路辅助",
        "DeepSeek API": "LLM 辅助写作（无密钥探测）",
        "HF Daily Papers": "按 JSON API 端点探测（清单记录的 RSS 端点此前 401/404）",
        "Berkeley RDI Blog": "清单未给完整域名，按 rdi.berkeley.edu 推测探测",
        "Substack 类 newsletter": "以 latent.space 为样本探测",
        "LMSYS Blog": "按 lmsys.org/feed.xml 探测",
    }

    wb = load_workbook(ROOT / "reports" / "source-inventory" / "信源清单-2026-08-28.xlsx")
    ws = wb["全部信源明细"]
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    BODY_FONT = Font(name="微软雅黑", size=10)
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    WRAP = Alignment(wrap_text=True, vertical="top")
    GREEN = PatternFill("solid", fgColor="C6EFCE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    RED = PatternFill("solid", fgColor="FFC7CE")
    GRAY = PatternFill("solid", fgColor="D9D9D9")

    new_headers = ["本地直连", "经本地代理(127.0.0.1:7897)", "结论与备注（2026-08-31 探测）"]
    for i, h in enumerate(new_headers):
        c = ws.cell(row=1, column=7 + i, value=h)
        c.fill, c.font, c.border, c.alignment = HEADER_FILL, HEADER_FONT, BORDER, WRAP

    for ridx in range(2, ws.max_row + 1):
        cat = str(ws.cell(row=ridx, column=1).value or "")
        name = str(ws.cell(row=ridx, column=2).value or "")
        d = p = note = None
        key = (cat, name)
        # round-1 built-in rows
        matched = False
        for sheet_name, probe_name in R1_MAP:
            if name == sheet_name:
                d = r1_text("direct", probe_name)
                p = r1_text("proxy", probe_name)
                matched = True
                break
        if not matched:
            if name == "Google AI Blog" and "内置" in cat:
                d, p = r1_text("direct", "Google AI Blog"), r1_text("proxy", "Google AI Blog")
            elif name == "Google AI Blog（示例）":
                d, p = feed_pair("Google AI Blog (OPML URL)")
            elif name == "OpenAI News（示例）":
                d, p = r1_text("direct", "OpenAI News"), r1_text("proxy", "OpenAI News")
            elif name == "Hugging Face Blog（示例）":
                d, p = r1_text("direct", "Hugging Face Blog"), r1_text("proxy", "Hugging Face Blog")
            elif name == "Google DeepMind Blog（示例）":
                d, p = r1_text("direct", "Google DeepMind"), r1_text("proxy", "Google DeepMind")
            elif name == "Microsoft AI Blog（示例）":
                d, p = feed_pair("Microsoft AI Blog")
            elif name == "Wired AI（示例）":
                d, p = feed_pair("Wired AI")
            elif name == "InfoQ CN（示例）":
                d, p = feed_pair("InfoQ CN")
            elif name == "NVIDIA Generative AI Blog（示例）":
                d, p = feed_pair("NVIDIA Generative AI Blog")
            elif name == "宝玉（示例）":
                d, p = feed_pair("Baoyu")
            elif name == "Simon Willison（示例）":
                d, p = feed_pair("Simon Willison")
            elif name == "Telegram 桥接":
                d, p = feed_pair("Telegram bridge (sample: durov)")
            elif name == "即刻桥接":
                d, p = feed_pair("Jike bridge (host)")
            elif name == "Karpathy X（候选）":
                d, p = feed_pair("Karpathy X (rsshub.pseudoyu.com)")
            elif name.startswith("X API"):
                d, p = feed_pair("X API")
            elif name.startswith("SocialData"):
                d, p = feed_pair("SocialData.tools")
            elif name.startswith("TikHub"):
                d, p = feed_pair("TikHub")
            elif name.startswith("AgentMail"):
                d, p = feed_pair("AgentMail")
            elif name == "Jina Reader":
                d, p = feed_pair("Jina Reader")
            elif name == "Google Translate":
                d, p = feed_pair("Google Translate")
            elif name == "DeepSeek API":
                d, p = feed_pair("DeepSeek API")
            elif name.startswith("OpenRouter"):
                d, p = feed_pair("OpenRouter Announcements")
            elif name.startswith("LMSYS"):
                d, p = feed_pair("LMSYS Blog")
            elif name.startswith("Hugging Face Daily Papers"):
                d, p = feed_pair("HF Daily Papers (JSON API)")
            elif name.startswith("Berkeley RDI"):
                d, p = feed_pair("Berkeley RDI Blog (guessed)")
            elif name.startswith("量子位"):
                d, p = feed_pair("QbitAI")
            elif name.startswith("Substack"):
                d, p = feed_pair("Substack (latent.space, sample)")
            else:
                d = p = "—"  # rule rows, local-file entry, skip routes
        note = NOTES.get(name, "")
        for col, val in ((7, d), (8, p), (9, note)):
            c = ws.cell(row=ridx, column=col, value=val)
            c.font, c.border, c.alignment = BODY_FONT, BORDER, WRAP
        fill = None
        if d and p:
            dok = d.startswith("✓")
            pok = p.startswith("✓")
            if d == "—" and p == "—":
                fill = GRAY
            elif dok:
                fill = GREEN
            elif pok:
                fill = YELLOW
            else:
                fill = RED
        for col in (7, 8):
            if fill is not None:
                ws.cell(row=ridx, column=col).fill = fill

    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 56

    out = ROOT / "reports" / "source-inventory" / "信源清单-本地网络标注-2026-08-31.xlsx"
    wb.save(out)
    print(f"\nsaved: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
