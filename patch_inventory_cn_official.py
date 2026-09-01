#!/usr/bin/env python3
"""Append the 2026-09-01 CN official source rows to the inventory xlsx.

Additive only: inserts 5 built-in official rows (DeepSeek, MiniMax,
ByteDance Seed, Zhipu AI, Moonshot/Kimi) after the last official row and
one Qwen evaluated-not-adopted row. No existing cell is modified.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

ROOT = Path(__file__).resolve().parent
PATH = ROOT / "reports" / "source-inventory" / "信源清单-本地网络标注-2026-08-31.xlsx"

COLORS = {
    "高级信源·默认关闭": "E4DFEC",
    "辅助服务·非信源": "EDEDED",
    "已评估未接入": "FCE4D6",
}
BODY_FONT = Font(name="微软雅黑", size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")


def last_category_row(ws, cat: str) -> int | None:
    last = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == cat:
            last = r
    return last


def style_row(ws, ridx: int, cat: str):
    fill = PatternFill("solid", fgColor=COLORS.get(cat, "FFFFFF"))
    for c in range(1, 10):
        cobj = ws.cell(row=ridx, column=c)
        cobj.font, cobj.border, cobj.alignment = BODY_FONT, BORDER, WRAP
        if c <= 6:
            cobj.fill = fill


def paint_status(ws, ridx: int):
    d = str(ws.cell(row=ridx, column=7).value or "")
    p = str(ws.cell(row=ridx, column=8).value or "")
    dok, pok = d.startswith("✓"), p.startswith("✓")
    ws.cell(row=ridx, column=7).fill = GREEN if dok else (YELLOW if pok else RED)
    if p == "—":
        return  # proxy not tested for this source; leave neutral
    ws.cell(row=ridx, column=8).fill = GREEN if pok else (YELLOW if dok else RED)


def write_row(ws, ridx: int, rowdata):
    for c, v in enumerate(rowdata, 1):
        ws.cell(row=ridx, column=c, value=v)
    style_row(ws, ridx, rowdata[0])
    paint_status(ws, ridx)


wb = load_workbook(PATH)
ws = wb["全部信源明细"]

# ---- 1. five new built-in official CN rows after the last official row ----
NEW_OFFICIAL = [
    ("内置·官方AI动态", "DeepSeek", "deepseek.com",
     "https://deepseek.com/news", "HTML 页面解析（Next.js flight payload）",
     "title/date/slug 内嵌 JSON；最多 10 条；保留窗口 45 天",
     "✓ HTTP 200", "—",
     "2026-09-01 接入；直连可用；最新一篇 2026-04-24 超出 45 天窗口，"
     "今日贡献 0 条（新公告进入窗口后自动收录）"),
    ("内置·官方AI动态", "MiniMax", "minimaxi.com",
     "https://www.minimaxi.com/api/news", "公开 JSON API",
     "publishDate 毫秒时间戳；最多 10 条",
     "✓ 2 items", "—",
     "2026-09-01 接入；直连可用，官方公开 JSON 接口"),
    ("内置·官方AI动态", "ByteDance Seed", "seed.bytedance.com",
     "https://seed.bytedance.com/blog", "HTML 页面解析（内嵌 article_list JSON）",
     "ArticleMeta ID/Status/PublishDate；仅 Status=2 已发布；最多 10 条",
     "✓ 3 items", "—",
     "2026-09-01 接入；直连可用；中文标题优先"),
    ("内置·官方AI动态", "智谱 AI", "zhipuai.cn",
     "https://www.zhipuai.cn/news", "Next.js RSC payload（RSC: 1 头）",
     "navConfig article id/title/createAt；最多 10 条",
     "✓ HTTP 200", "—",
     "2026-09-01 接入；直连可用；最新一篇 2026-06-16 超出 45 天窗口，今日贡献 0 条"),
    ("内置·官方AI动态", "月之暗面 Moonshot（Kimi）", "kimi.com",
     "https://www.kimi.com/blog/", "HTML 页面解析（卡片 aria-label + card-date）",
     "moonshot.cn 研究/新闻路由至国际站 kimi.com；最多 10 条",
     "✓ HTTP 200", "—",
     "2026-09-01 接入；直连可用；最新一篇 2026-07-16 超出 45 天窗口，今日贡献 0 条"),
]
anchor = last_category_row(ws, "内置·官方AI动态")
if anchor is None:
    anchor = ws.max_row
ws.insert_rows(anchor + 1, amount=len(NEW_OFFICIAL))
for i, rowdata in enumerate(NEW_OFFICIAL):
    write_row(ws, anchor + 1 + i, rowdata)

# ---- 2. Qwen evaluated-not-adopted row (after last evaluated row, else end) ----
QWEN_ROW = (
    "已评估未接入", "通义千问 Qwen", "qwen.ai",
    "https://qwen.ai（纯 SPA）", "无公开 RSS/API",
    "2026-09-01 评估：/api/v2/article 重定向到内网 :8080 主机，外网不可达；"
    "无 sitemap/robots 线索；openapi.json 仅暴露内部测试路由",
    "✗ 无公开抓取路径", "—",
    "不接入：私有 API 不稳定且不可外网访问；待阿里官方公开 feed 再议")
qanchor = last_category_row(ws, "已评估未接入")
if qanchor is None:
    write_row(ws, ws.max_row + 1, QWEN_ROW)
else:
    ws.insert_rows(qanchor + 1, amount=1)
    write_row(ws, qanchor + 1, QWEN_ROW)

# keep the filter covering the new rows (only extends the range)
ws.auto_filter.ref = f"A1:I{ws.max_row}"

wb.save(PATH)
print("patched:", PATH, "| detail rows:", ws.max_row - 1)
