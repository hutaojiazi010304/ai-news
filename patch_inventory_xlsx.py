#!/usr/bin/env python3
"""Patch the annotated inventory xlsx: fix WaytoAGI count and insert the
11 rows missing from the outdated 2026-08-28 sheet (advanced sources,
DeepSeek, evaluated-not-adopted) at their proper category positions,
using round-2 probe results."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

ROOT = Path(__file__).resolve().parent
PATH = ROOT / "reports" / "source-inventory" / "信源清单-本地网络标注-2026-08-31.xlsx"

r2 = {}
for row in json.loads((ROOT / "probe-inventory-report.json").read_text(encoding="utf-8")):
    r2[(row["mode"], row["label"])] = row


def cell(label: str, mode: str) -> str:
    row = r2.get((mode, label))
    if not row:
        return "—"
    if row["kind"] == "feed":
        if row["ok"]:
            return f"✓ {row['entries']} entries"
        if row["status"]:
            return f"✗ HTTP {row['status']}"
        return f"✗ {row['error'].split(':')[0]}"
    if row["ok"]:
        return f"✓ HTTP {row['status']}"
    if row["status"]:
        return f"✗ HTTP {row['status']}"
    return f"✗ {row['error'].split(':')[0]}"


COLORS = {
    "高级信源·默认关闭": "E4DFEC",
    "辅助服务·非信源": "EDEDED",
    "已评估未接入": "FCE4D6",
}
BODY_FONT = Font(name="微软雅黑", size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")

ADVANCED_ROWS = [
    ("高级信源·默认关闭", "X API 官方近期搜索", "api.x.com",
     "X API v2 recent search", "官方 X API（付费）",
     "X_API_ENABLED=1 + X_BEARER_TOKEN；无密钥探测，401 = 网络可达",
     "X API", "直连被墙；走代理可达（401 为缺密钥的正常返回）"),
    ("高级信源·默认关闭", "SocialData.tools X 搜索 + 列表", "api.socialdata.tools",
     "关键词搜索 + 精选列表", "第三方 X 数据 API（付费）",
     "SOCIALDATA_ENABLED=1 + SOCIALDATA_API_KEY",
     "SocialData.tools", "网络可达（404 为路径不匹配，主机连通）"),
    ("高级信源·默认关闭", "TikHub 抖音 / 小红书搜索", "api.tikhub.io",
     "抖音 / 小红书搜索接口", "付费聚合 API",
     "TIKHUB_ENABLED=1 + TIKHUB_API_KEY",
     "TikHub", "网络可达，无需代理"),
    ("高级信源·默认关闭", "AgentMail 邮件摘要", "api.agentmail.to",
     "list-messages 元数据接口", "AgentMail API（付费）",
     "EMAIL_DIGEST_ENABLED=1 + AGENTMAIL_API_KEY + AGENTMAIL_INBOX_ID",
     "AgentMail", "网络可达，无需代理"),
]
DEEPSEEK_ROW = (
    "辅助服务·非信源", "DeepSeek API", "api.deepseek.com",
    "LLM 接口", "LLM 辅助写作",
    "文章生成链路使用，非新闻信源",
    "DeepSeek API", "网络可达（401 为缺密钥的正常返回）",
)
EVALUATED_ROWS = [
    ("已评估未接入", "OpenRouter Announcements", "openrouter.ai",
     "https://openrouter.ai/blog/feed.xml", "RSS（未接入）",
     "2026-05 记录：接入探测时返回 Not Found 页面",
     "OpenRouter Announcements", "本次探测恢复可用（116 条），可重新评估接入"),
    ("已评估未接入", "LMSYS Blog", "lmsys.org",
     "https://lmsys.org/feed.xml", "RSS（未接入）",
     "重定向或 404",
     "LMSYS Blog", "本次探测确认 404"),
    ("已评估未接入", "Hugging Face Daily Papers", "huggingface.co",
     "https://huggingface.co/api/daily_papers", "JSON API（未接入）",
     "RSS 风格端点返回 401/404；等稳定公开 feed 再议",
     "HF Daily Papers (JSON API)", "JSON API 端点走代理可用（289KB）；huggingface.co 直连不通"),
    ("已评估未接入", "Berkeley RDI Blog", "rdi.berkeley.edu",
     "https://rdi.berkeley.edu/feed.xml", "RSS（未接入）",
     "仅 2021 年 Jekyll 占位条目；清单未给完整域名，按此推测探测",
     "Berkeley RDI Blog (guessed)", "可达，但仍只有 1 条占位内容，无接入价值"),
    ("已评估未接入", "量子位 QbitAI", "qbitai.com",
     "https://www.qbitai.com/feed", "RSS（观察名单）",
     "项目抓取路径下返回 403",
     "QbitAI", "本次直连可用（10 条，浏览器 UA）；走代理反而超时——与旧记录相反，接入前建议用管线 UA 复核"),
    ("已评估未接入", "Substack 类 newsletter", "substack.com",
     "各刊 RSS（样本：https://www.latent.space/feed）", "RSS（未入公开示例）",
     "GitHub Actions 常收 403，本地可能正常",
     "Substack (latent.space, sample)", "样本直连可用（20 条），印证「本地可能正常」"),
]


def write_row(ws, ridx: int, rowdata):
    cat, name, domain, url, method, note_base, label, extra = rowdata
    fill = PatternFill("solid", fgColor=COLORS.get(cat, "FFFFFF"))
    d, p = cell(label, "direct"), cell(label, "proxy")
    values = [cat, name, domain, url, method, note_base, d, p, extra]
    for c, v in enumerate(values, 1):
        cobj = ws.cell(row=ridx, column=c, value=v)
        cobj.font, cobj.border, cobj.alignment = BODY_FONT, BORDER, WRAP
        if c <= 6:
            cobj.fill = fill
    dok, pok = d.startswith("✓"), p.startswith("✓")
    gfill = GREEN if dok else (YELLOW if pok else RED)
    ws.cell(row=ridx, column=7).fill = gfill
    ws.cell(row=ridx, column=8).fill = gfill


def find_row(ws, name: str) -> int:
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == name:
            return r
    raise KeyError(name)


wb = load_workbook(PATH)
ws = wb["全部信源明细"]

# fix WaytoAGI wrong count from round-1 probe (payload key mismatch)
wrow = find_row(ws, "WaytoAGI")
ws.cell(row=wrow, column=7).value = "✓ 35 items"
ws.cell(row=wrow, column=8).value = "✓ 35 items"

# insert advanced block before the 辅助服务 section (Jina Reader row)
anchor = find_row(ws, "Jina Reader")
ws.insert_rows(anchor, amount=len(ADVANCED_ROWS))
for i, rowdata in enumerate(ADVANCED_ROWS):
    write_row(ws, anchor + i, rowdata)

# insert DeepSeek after Google Translate
gt = find_row(ws, "Google Translate")
ws.insert_rows(gt + 1, amount=1)
write_row(ws, gt + 1, DEEPSEEK_ROW)

# append evaluated-not-adopted rows at the end
ridx = ws.max_row
for rowdata in EVALUATED_ROWS:
    ridx += 1
    write_row(ws, ridx, rowdata)

ws.auto_filter.ref = f"A1:I{ws.max_row}"
wb.save(PATH)
print("patched:", PATH, "| total rows:", ws.max_row - 1)
