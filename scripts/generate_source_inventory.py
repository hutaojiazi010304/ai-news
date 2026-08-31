"""Generate a readable Excel inventory of all AI News Radar sources.

One-off report helper; output goes to reports/.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = ["分类", "名称", "域名", "具体地址 / URL", "接入方式", "备注 / 过滤规则"]

# (分类, 名称, 域名, 地址, 接入方式, 备注)
ROWS = [
    # ---------------- 内置 · 官方 AI 动态 ----------------
    ("内置·官方AI动态", "OpenAI News", "openai.com",
     "https://openai.com/news/rss.xml", "官方 RSS", "保留窗口 45 天"),
    ("内置·官方AI动态", "Google DeepMind", "deepmind.google",
     "https://deepmind.google/blog/rss.xml", "官方 RSS", "保留窗口 45 天"),
    ("内置·官方AI动态", "Google AI Blog", "blog.google",
     "https://blog.google/innovation-and-ai/technology/ai/rss/", "官方 RSS", "保留窗口 45 天"),
    ("内置·官方AI动态", "Hugging Face Blog", "huggingface.co",
     "https://huggingface.co/blog/feed.xml", "官方 RSS", "保留窗口 45 天"),
    ("内置·官方AI动态", "GitHub AI & ML", "github.blog",
     "https://github.blog/ai-and-ml/feed/", "官方 RSS", "保留窗口 45 天"),
    ("内置·官方AI动态", "OpenAI Skills", "github.com",
     "https://github.com/openai/skills/commits/main.atom", "GitHub commit Atom",
     "关键词过滤：hatch / pet / migrate-to-codex"),
    ("内置·官方AI动态", "Anthropic News", "anthropic.com",
     "https://www.anthropic.com/news", "HTML 页面解析",
     "解析 /news/ 链接 + <time>；无官方 RSS"),
    ("内置·官方AI动态", "OpenAI Codex Changelog", "developers.openai.com",
     "https://developers.openai.com/codex/changelog", "HTML 页面解析",
     "解析 li[id] + time/h3"),

    # ---------------- 内置 · 精选 AI 媒体 ----------------
    ("内置·精选媒体", "The Decoder AI News", "the-decoder.com",
     "https://the-decoder.com/feed/", "RSS", "每源每次最多 10 条；保留窗口 30 天"),
    ("内置·精选媒体", "TechCrunch AI", "techcrunch.com",
     "https://techcrunch.com/category/artificial-intelligence/feed/", "RSS 分类 feed", "最多 8 条"),
    ("内置·精选媒体", "The Verge", "theverge.com",
     "https://www.theverge.com/rss/index.xml", "全站 RSS + 严格标题 AI 关键词过滤",
     "最多 6 条；AI 专题 RSS 不稳定，故用全站 feed 过滤"),
    ("内置·精选媒体", "MarkTechPost Research", "marktechpost.com",
     "https://www.marktechpost.com/feed/", "RSS + 研究关键词过滤",
     "research_only，走研究泳道并降权；最多 6 条"),
    ("内置·精选媒体", "VentureBeat AI", "venturebeat.com",
     "https://venturebeat.com/category/ai/feed", "RSS 分类 feed", "最多 8 条"),
    ("内置·精选媒体", "Artificial Intelligence News", "artificialintelligence-news.com",
     "https://www.artificialintelligence-news.com/feed/", "RSS", "最多 8 条"),
    ("内置·精选媒体", "Claude Code Releases", "github.com",
     "https://github.com/anthropics/claude-code/releases.atom", "GitHub releases Atom", "最多 6 条"),

    # ---------------- 内置 · 聚合 / 社区 ----------------
    ("内置·聚合社区", "AI HOT", "aihot.virxact.com",
     "https://aihot.virxact.com/api/public/items", "公开 JSON API",
     "mode=selected；仅保留评分≥60；游标翻页 ≤5 页×100 条；RSS 仅兜底不默认使用"),
    ("内置·聚合社区", "AI Breakfast", "aibreakfast.beehiiv.com",
     "https://r.jina.ai/https://aibreakfast.beehiiv.com/", "Beehiiv 公开归档页（经 Jina Reader）",
     "Beehiiv 直连对 GitHub Actions 返回拦截，故走 r.jina.ai；正则解析归档 Markdown"),
    ("内置·聚合社区", "Follow Builders", "raw.githubusercontent.com",
     "…/zarazhangrui/follow-builders/main/feed-x.json、feed-blogs.json、feed-podcasts.json",
     "公开生成 JSON（该项目 Actions 用官方 X API 生成）",
     "本项目只读公开 feed 文件，无需 X 凭据；覆盖 X 推文/博客/播客"),
    ("内置·聚合社区", "AI HubToday", "hex2077.dev",
     "https://hex2077.dev/rss-zh-CN.xml", "RSS",
     "原 ai.hubtoday.app 迁移到 hex2077.dev 后改读结构化 RSS"),
    ("内置·聚合社区", "AIbase", "aibase.com",
     "https://www.aibase.com/zh/news", "HTML 页面解析", "a[href^=/news/] + h3 + 时间文本"),
    ("内置·聚合社区", "TechURLs", "techurls.com",
     "https://techurls.com/", "HTML 页面解析", "div.publisher-block 结构解析"),
    ("内置·聚合社区", "Buzzing", "buzzing.cc",
     "https://www.buzzing.cc/feed.json", "JSON Feed", "讨论层，抓取上限 50 条"),
    ("内置·聚合社区", "Info Flow (Iris)", "iris.findtruman.io",
     "https://iris.findtruman.io/web/info_flow", "页面提取内嵌 feed 列表 → 逐个解析 RSS",
     "讨论层，抓取上限 50 条；单个子 feed 失败自动跳过"),
    ("内置·聚合社区", "BestBlogs", "bestblogs.dev",
     "https://api.bestblogs.dev/api/newsletter/list", "POST JSON API（分页 ≤12 页）",
     "失败时回退抓 www.bestblogs.dev/en/newsletter 页面"),
    ("内置·聚合社区", "Zeli（HN 24h 最热）", "zeli.app",
     "https://zeli.app/api/hacker-news?type=hot24h", "公开 JSON API", "Hacker News 24 小时最热帖"),
    ("内置·聚合社区", "Hacker News Algolia", "hn.algolia.com",
     "https://hn.algolia.com/api/v1/search_by_date", "公开 Algolia 搜索 API",
     "18 组 AI 关键词查询、24h 窗口；要求多关键词命中且 num_comments≥2 或 points≥10"),
    ("内置·聚合社区", "NewsNow", "newsnow.busiyi.world",
     "首页 → JS bundle 提取源列表 → POST /api/s/entire（失败回退 GET /api/s?id=）",
     "页面 + API 混合抓取",
     "覆盖 hackernews、producthunt、github、sspai、juejin、36kr 等约 57 个子源"),
    ("内置·聚合社区", "WaytoAGI", "waytoagi.feishu.cn",
     "https://waytoagi.feishu.cn/wiki/QPe5w5g7UisbEkkow8XcDmOpn8e（历史页兜底）",
     "飞书 Wiki 页面解析（block_map）",
     "产出 data/waytoagi-7d.json；最新日条目进入 Community 频道"),

    # ---------------- OPML 订阅层 ----------------
    ("OPML订阅层", "OPML 私有订阅（入口）", "—",
     "本地：feeds/follow.opml（不入库）；Actions：FOLLOW_OPML_B64 secret",
     "OPML 导入 RSS/Atom",
     "secret 未配置时回退公开示例 feeds/follow.example.opml"),
    ("OPML订阅层", "OpenAI News（示例）", "openai.com",
     "https://openai.com/news/rss.xml", "RSS（示例 OPML）", "feeds/follow.example.opml"),
    ("OPML订阅层", "Hugging Face Blog（示例）", "huggingface.co",
     "https://huggingface.co/blog/feed.xml", "RSS（示例 OPML）", "feeds/follow.example.opml"),
    ("OPML订阅层", "Google DeepMind Blog（示例）", "deepmind.google",
     "https://deepmind.google/blog/rss.xml", "RSS（示例 OPML）", "feeds/follow.example.opml"),
    ("OPML订阅层", "Google AI Blog（示例）", "blog.google",
     "https://blog.google/technology/ai/rss/", "RSS（示例 OPML）", "feeds/follow.example.opml"),
    ("OPML订阅层", "Microsoft AI Blog（示例）", "news.microsoft.com",
     "https://news.microsoft.com/source/topics/ai/feed/", "RSS（示例 OPML）", "feeds/follow.example.opml"),
    ("OPML订阅层", "Wired AI（示例）", "wired.com",
     "https://www.wired.com/feed/tag/ai/latest/rss", "RSS（示例 OPML）", "feeds/follow.example.opml"),
    ("OPML订阅层", "InfoQ CN（示例）", "infoq.cn",
     "https://www.infoq.cn/feed", "RSS（示例 OPML）", "feeds/follow.example.opml"),
    ("OPML订阅层", "NVIDIA Generative AI Blog（示例）", "developer.nvidia.com",
     "https://developer.nvidia.com/blog/category/generative-ai/feed/", "RSS（示例 OPML）",
     "feeds/follow.example.opml"),
    ("OPML订阅层", "宝玉（示例）", "baoyu.io",
     "https://baoyu.io/feed.xml", "RSS（示例 OPML）", "feeds/follow.example.opml"),
    ("OPML订阅层", "Simon Willison（示例）", "simonwillison.net",
     "https://simonwillison.net/atom/everything/", "Atom（示例 OPML）", "feeds/follow.example.opml"),
    ("OPML订阅层", "RSSHub 路由替换规则", "rsshub.app → 官方域名",
     "infoq / huggingface blog-zh / readhub / 36kr / sspai / 美团技术",
     "RSSHub 路由自动替换为官方 feed",
     "如 rsshub.app/infoq/recommend → www.infoq.cn/feed"),
    ("OPML订阅层", "OPML 跳过路由", "rsshub.app 等",
     "telegram、jike、bilibili、zhihu、小宇宙播客、xyzrank、MIT科技评论中文、wechat2rss、werss 等",
     "直接跳过", "桥接不稳定或无官方 RSS，默认跳过"),
    ("OPML订阅层", "Telegram 桥接", "t.me",
     "https://t.me/s/<频道名>", "直抓 Telegram 公开页",
     "RSSHub telegram/channel 路由自动改走公开页解析"),
    ("OPML订阅层", "即刻桥接", "m.okjike.com",
     "https://m.okjike.com/topics/<id> 或 /users/<id>", "直抓即刻公开页（__NEXT_DATA__）",
     "RSSHub jike 路由自动改走公开页解析"),
    ("OPML订阅层", "Karpathy X（候选）", "rsshub.pseudoyu.com",
     "https://rsshub.pseudoyu.com/twitter/user/karpathy", "RSSHub X 路由",
     "仅示例（feeds/social-x.example.opml）；建议私有 OPML 使用，公共实例不稳定"),

    # ---------------- 高级信源（默认关闭） ----------------
    ("高级信源·默认关闭", "X API 官方近期搜索", "api.x.com",
     "X API v2 recent search", "官方 X API（付费）",
     "X_API_ENABLED=1 + X_BEARER_TOKEN；X_API_MAX_RESULTS / 每日上限 / 运行时段控制"),
    ("高级信源·默认关闭", "SocialData.tools X 搜索 + 列表", "api.socialdata.tools",
     "关键词搜索（游标翻页）+ 精选列表「AI is cool, i guess」（@aiwarts，list id 1695376776867062037）",
     "第三方 X 数据 API（付费）",
     "SOCIALDATA_ENABLED=1 + SOCIALDATA_API_KEY；仅原创帖、4 天窗口、列表 ≤10 页"),
    ("高级信源·默认关闭", "TikHub 抖音 / 小红书搜索", "api.tikhub.io",
     "抖音：douyin_search.fetch_general_search_v2；小红书：xiaohongshu_app_v2.search_notes（回退 web_v3）",
     "付费聚合 API",
     "TIKHUB_ENABLED=1 + TIKHUB_API_KEY；最多点赞排序、一周内；代码内强制 7 天窗口；默认关键词 OpenAI,Claude,大模型,Agent,AI工具…"),
    ("高级信源·默认关闭", "AgentMail 邮件摘要", "api.agentmail.to",
     "list-messages 元数据接口（不读正文 / 原始邮件）", "AgentMail API（付费）",
     "EMAIL_DIGEST_ENABLED=1 + AGENTMAIL_API_KEY + AGENTMAIL_INBOX_ID；发布到 Pages 需 EMAIL_DIGEST_PUBLISH=1"),

    # ---------------- 辅助服务 ----------------
    ("辅助服务·非信源", "Jina Reader", "r.jina.ai",
     "https://r.jina.ai/<目标页>", "网页转 Markdown 读取服务",
     "用于 AI Breakfast 抓取与条目标题/正文上下文兜底"),
    ("辅助服务·非信源", "Google Translate", "translate.googleapis.com",
     "翻译接口", "翻译支持", "文章生成链路使用，非新闻信源"),
    ("辅助服务·非信源", "DeepSeek API", "api.deepseek.com",
     "LLM 接口", "LLM 辅助写作", "文章生成链路使用，非新闻信源"),

    # ---------------- 已评估未接入 ----------------
    ("已评估未接入", "OpenRouter Announcements", "openrouter.ai",
     "/blog/feed.xml", "RSS（未接入）", "接入探测时返回 Not Found 页面"),
    ("已评估未接入", "LMSYS Blog", "lmsys.org",
     "feed 端点探测", "RSS（未接入）", "重定向或 404"),
    ("已评估未接入", "Hugging Face Daily Papers", "huggingface.co",
     "RSS 风格端点探测", "RSS（未接入）", "返回 401/404；等稳定公开 feed 再议"),
    ("已评估未接入", "Berkeley RDI Blog", "—",
     "/feed.xml", "RSS（未接入）", "仅 2021 年 Jekyll 占位条目"),
    ("已评估未接入", "量子位 QbitAI", "qbitai.com",
     "RSS 直连", "RSS（观察名单）", "项目抓取路径下返回 403"),
    ("已评估未接入", "Substack 类 newsletter", "substack.com",
     "各刊 RSS", "RSS（未入公开示例）", "GitHub Actions 常收 403，本地可能正常"),
]

CATEGORY_COLORS = {
    "内置·官方AI动态": "DCE6F1",
    "内置·精选媒体": "E2EFDA",
    "内置·聚合社区": "FFF2CC",
    "OPML订阅层": "F2DCDB",
    "高级信源·默认关闭": "E4DFEC",
    "辅助服务·非信源": "EDEDED",
    "已评估未接入": "FCE4D6",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="微软雅黑", size=10)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

wb = Workbook()

# ---------- Sheet 1: 汇总 ----------
ws_sum = wb.active
ws_sum.title = "汇总"
ws_sum["A1"] = "AI News Radar 信源清单"
ws_sum["A1"].font = TITLE_FONT
ws_sum["A2"] = "生成日期：2026-08-28　|　来源：scripts/update_news.py、feeds/*.opml、docs/SOURCE_COVERAGE.md"
ws_sum["A2"].font = Font(name="微软雅黑", size=9, color="808080")

sum_headers = ["分类", "数量", "说明"]
ws_sum.append([])
ws_sum.append(sum_headers)
for col, _ in enumerate(sum_headers, 1):
    cell = ws_sum.cell(row=4, column=col)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
    cell.alignment = WRAP_CENTER

summary_rows = [
    ("内置·官方AI动态", "官方 RSS / Atom / 页面解析，保留 45 天窗口；全部公开免密钥"),
    ("内置·精选媒体", "公开 RSS/Atom + 每源条数上限与关键词过滤，保留 30 天窗口"),
    ("内置·聚合社区", "公开 JSON API / JSON Feed / HTML 解析 / 第三方聚合，免费公开路径"),
    ("OPML订阅层", "私有 feeds/follow.opml（FOLLOW_OPML_B64）；含公开示例 10 条与路由替换/桥接规则"),
    ("高级信源·默认关闭", "需环境变量 + 密钥开启：X API、SocialData、TikHub、AgentMail"),
    ("辅助服务·非信源", "抓取/写作链路辅助服务，不是新闻来源"),
    ("已评估未接入", "探测失败、403 或低价值，暂不进入默认集"),
]
counts: dict[str, int] = {}
for row in ROWS:
    counts[row[0]] = counts.get(row[0], 0) + 1
r = 5
for cat, desc in summary_rows:
    ws_sum.cell(row=r, column=1, value=cat)
    ws_sum.cell(row=r, column=2, value=counts.get(cat, 0))
    ws_sum.cell(row=r, column=3, value=desc)
    fill = PatternFill("solid", fgColor=CATEGORY_COLORS.get(cat, "FFFFFF"))
    for c in range(1, 4):
        cell = ws_sum.cell(row=r, column=c)
        cell.font = BODY_FONT
        cell.border = BORDER
        cell.alignment = WRAP
        cell.fill = fill
    r += 1
total = ws_sum.cell(row=r, column=1, value="合计")
total.font = Font(name="微软雅黑", bold=True, size=10)
ws_sum.cell(row=r, column=2, value=len(ROWS)).font = Font(name="微软雅黑", bold=True, size=10)
for c in range(1, 4):
    ws_sum.cell(row=r, column=c).border = BORDER
ws_sum.column_dimensions["A"].width = 22
ws_sum.column_dimensions["B"].width = 8
ws_sum.column_dimensions["C"].width = 78

# ---------- Sheet 2: 全部信源明细 ----------
ws = wb.create_sheet("全部信源明细")
ws.append(HEADERS)
for col in range(1, len(HEADERS) + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
    cell.alignment = WRAP_CENTER

for row in ROWS:
    ws.append(row)

for r in range(2, ws.max_row + 1):
    cat = ws.cell(row=r, column=1).value
    fill = PatternFill("solid", fgColor=CATEGORY_COLORS.get(cat, "FFFFFF"))
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = BODY_FONT
        cell.border = BORDER
        cell.alignment = WRAP
        cell.fill = fill

widths = [18, 26, 28, 56, 30, 52]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:F{ws.max_row}"

# ---------- Sheet 3: 接入方式速查 ----------
ws3 = wb.create_sheet("接入方式速查")
ws3.append(["接入方式", "说明", "代表信源"])
for col in range(1, 4):
    cell = ws3.cell(row=1, column=col)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
methods = [
    ("官方 RSS / Atom", "最优先的默认接入方式，稳定且免密钥",
     "OpenAI News、DeepMind、Google AI、Hugging Face、GitHub AI & ML、The Decoder、TechCrunch AI 等"),
    ("GitHub Atom（commits/releases）", "GitHub 原生 Atom，用于仓库动态与版本发布",
     "OpenAI Skills（commits）、Claude Code Releases（releases）"),
    ("HTML 页面解析", "无 RSS 的稳定公开页面，用 requests + BeautifulSoup 解析",
     "Anthropic News、OpenAI Codex Changelog、AIbase、TechURLs、WaytoAGI 飞书 Wiki"),
    ("公开 JSON API", "带时间戳的稳定公开接口，自定义 fetcher",
     "AI HOT、Zeli、BestBlogs、Hacker News Algolia、NewsNow"),
    ("JSON Feed", "JSON Feed 标准格式", "Buzzing"),
    ("公开生成 feed 文件", "读取其他项目 Actions 生成的公开 JSON/RSS 文件",
     "Follow Builders（feed-x/blogs/podcasts.json）、AI HubToday（hex2077.dev RSS）"),
    ("Jina Reader 兜底", "原始站点拦截 Actions 时，经 r.jina.ai 读取", "AI Breakfast"),
    ("OPML 导入", "私有订阅入口：本地 --rss-opml 或 Actions FOLLOW_OPML_B64",
     "feeds/follow.example.opml 示例 10 条"),
    ("桥接直读公开页", "RSSHub 路由自动改抓平台公开页", "Telegram（t.me/s/）、即刻（m.okjike.com）"),
    ("付费 API 适配器（默认关闭）", "密钥 + 环境变量开启，含频次/条数上限",
     "X API、SocialData.tools、TikHub、AgentMail"),
]
for m in methods:
    ws3.append(m)
for r in range(2, ws3.max_row + 1):
    for c in range(1, 4):
        cell = ws3.cell(row=r, column=c)
        cell.font = BODY_FONT
        cell.border = BORDER
        cell.alignment = WRAP
ws3.column_dimensions["A"].width = 26
ws3.column_dimensions["B"].width = 52
ws3.column_dimensions["C"].width = 70
ws3.freeze_panes = "A2"

out = r"c:\Users\admin\Desktop\ai-news-radar-master\reports\source-inventory\信源清单-2026-08-28.xlsx"
import os
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print("saved:", out, "| rows:", len(ROWS))
