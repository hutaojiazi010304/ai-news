#!/usr/bin/env python3
"""Patch the annotated inventory xlsx with the latest 2026-08-31 results.

- OpenRouter / 量子位 / 新智元 adopted into 内置·精选媒体 (removed from 已评估未接入)
- 机器之心 added to 已评估未接入 (paid-service landing page, no free path)
- Status flips: Google DeepMind / Google AI Blog (direct resets),
  Info Flow Iris (502 upstream), AI HOT (proxy SSL error)
- Fresh counts: Follow Builders 23, Zeli 70
- DISABLED_SOURCES local-disable notes on all 9 affected rows
- 汇总 sheet re-synced (curated 7→10, 辅助 2→3, new rows for
  高级信源/已评估未接入, SUM range widened)
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

ROOT = Path(__file__).resolve().parent
PATH = ROOT / "reports" / "source-inventory" / "信源清单-本地网络标注-2026-08-31.xlsx"

COLORS = {
    "高级信源·默认关闭": "E4DFEC",
    "辅助服务·非信源": "EDEDED",
    "已评估未接入": "FCE4D6",
}
BODY_FONT = Font(name="微软雅黑", size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")


def find_row(ws, name: str, col: int = 2) -> int:
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=col).value == name:
            return r
    raise KeyError(name)


def style_row(ws, ridx: int, cat: str):
    fill = PatternFill("solid", fgColor=COLORS.get(cat, "FFFFFF"))
    for c in range(1, 10):
        cobj = ws.cell(row=ridx, column=c)
        cobj.font, cobj.border, cobj.alignment = BODY_FONT, BORDER, WRAP
        if c <= 6:
            cobj.fill = fill


def paint_status(ws, ridx: int):
    d = str(ws.cell(row=ridx, column=7).value or "")
    p = str(ws.cell(row=ridx, column=8).value or "")
    dok, pok = d.startswith("✓"), p.startswith("✓")
    fill = GREEN if dok else (YELLOW if pok else RED)
    ws.cell(row=ridx, column=7).fill = fill
    ws.cell(row=ridx, column=8).fill = fill


def write_row(ws, ridx: int, rowdata):
    cat, name, domain, url, method, note, direct, proxy, extra = rowdata
    values = [cat, name, domain, url, method, note, direct, proxy, extra]
    for c, v in enumerate(values, 1):
        ws.cell(row=ridx, column=c, value=v)
    style_row(ws, ridx, cat)
    paint_status(ws, ridx)


def append_note(ws, name: str, extra: str):
    r = find_row(ws, name)
    old = ws.cell(row=r, column=9).value
    ws.cell(row=r, column=9, value=f"{old}；{extra}" if old else extra)


wb = load_workbook(PATH)
ws = wb["全部信源明细"]

# ---- 1. drop the two evaluated rows that are now adopted ----
for name in ("量子位 QbitAI", "OpenRouter Announcements"):
    ws.delete_rows(find_row(ws, name))

# ---- 2. insert the three adopted feeds after Claude Code Releases ----
ADOPTED = [
    ("内置·精选媒体", "OpenRouter Announcements", "openrouter.ai",
     "https://openrouter.ai/blog/feed.xml", "RSS",
     "最多 6 条（平台动态类 feed，更新量小）；可信源关键词放行",
     "✓ 6 items", "✓ 6 items",
     "2026-08-31 接入为内置精选媒体（旧「Not Found」记录已失效）"),
    ("内置·精选媒体", "量子位（QbitAI）", "qbitai.com",
     "https://www.qbitai.com/feed", "RSS",
     "最多 10 条；拒绝旧版浏览器 UA（Chrome <126 返回 403），BROWSER_UA 已升级",
     "✓ 10 items", "✓ 10 items",
     "2026-08-31 接入为内置精选媒体；直连与代理均可用，加入可信源关键词"),
    ("内置·精选媒体", "新智元（AI Era）", "aiera.com.cn",
     "https://aiera.com.cn/feed/", "RSS（WordPress）",
     "最多 15 条（每晨一批中文 AI 新闻）；可信源关键词放行",
     "✓ 15 items", "✓ 15 items",
     "2026-08-31 接入为内置精选媒体；与近 7 日存档重叠 0%"),
]
anchor = find_row(ws, "Claude Code Releases") + 1
ws.insert_rows(anchor, amount=len(ADOPTED))
for i, rowdata in enumerate(ADOPTED):
    write_row(ws, anchor + i, rowdata)

# ---- 3. append 机器之心 to 已评估未接入 ----
write_row(ws, ws.max_row + 1, (
    "已评估未接入", "机器之心（Synced）", "jiqizhixin.com",
    "https://www.jiqizhixin.com/（全站路由均返回同一落地页）",
    "无免费 RSS/API",
    "2026-08-31 评估：整站已改为付费数据服务落地页；sitemap 仍更新但文章页与 /rss 均失效",
    "✗ 无免费 feed", "✗ 无免费 feed",
    "不接入：仅剩付费 API（需密钥）或微信公众号桥接（维护风险高）两条路"))

# ---- 4. in-place status updates ----
def set_direct(ws, name, g, note):
    r = find_row(ws, name)
    ws.cell(row=r, column=7, value=g)
    ws.cell(row=r, column=9, value=note)
    paint_status(ws, r)

set_direct(ws, "Google DeepMind", "✗ 连接被重置",
           "2026-08-31 直连多次复测均连接重置（当日早间曾通过）；已通过 DISABLED_SOURCES 本地关闭")
set_direct(ws, "Google AI Blog", "✗ 连接被重置",
           "2026-08-31 直连连接重置；已通过 DISABLED_SOURCES 本地关闭")
set_direct(ws, "Info Flow (Iris)", "✗ HTTP 502",
           "早间可用（50 条，约 46s），午后复测 3 次全部 502（上游故障）；已本地关闭，复测后再启用")
set_direct(ws, "Google DeepMind Blog（示例）", "✗ 连接被重置",
           "直连连接重置；已通过 DISABLED_SOURCES 本地关闭")
set_direct(ws, "Google AI Blog（示例）", "✗ 连接被重置",
           "直连连接重置；已通过 DISABLED_SOURCES 本地关闭")

r = find_row(ws, "AI HOT")
ws.cell(row=r, column=8, value="✗ SSLError")
ws.cell(row=r, column=9, value="直连正常；本地代理反而 SSL 握手失败")
paint_status(ws, r)

r = find_row(ws, "Follow Builders")
ws.cell(row=r, column=7, value="✓ 23 items")
ws.cell(row=r, column=8, value="✓ 23 items")
paint_status(ws, r)

r = find_row(ws, "Zeli（HN 24h 最热）")
ws.cell(row=r, column=7, value="✓ 70 items")
ws.cell(row=r, column=8, value="✓ 70 items")
paint_status(ws, r)

# ---- 5. DISABLED_SOURCES notes on the locally-disabled rows ----
DISABLE_NOTES = [
    ("Hugging Face Blog", "已本地关闭（DISABLED_SOURCES）"),
    ("OpenAI Skills", "已本地关闭（DISABLED_SOURCES）"),
    ("Claude Code Releases", "已本地关闭（DISABLED_SOURCES）"),
    ("AI Breakfast", "已本地关闭（DISABLED_SOURCES）"),
    ("NewsNow", "已本地关闭（DISABLED_SOURCES，云端 Actions 正常）"),
    ("Hugging Face Blog（示例）", "直连超时；已本地关闭（DISABLED_SOURCES）"),
    ("Microsoft AI Blog（示例）", "已本地关闭（DISABLED_SOURCES）"),
]
for name, extra in DISABLE_NOTES:
    append_note(ws, name, extra)

ws.auto_filter.ref = f"A1:I{ws.max_row}"

# ---- 6. 汇总 sheet ----
s = wb["汇总"]
s.cell(row=2, column=1, value=(
    "盘点基线：2026-08-28｜本地网络标注与接入更新：2026-08-31｜"
    "来源：scripts/update_news.py、feeds/*.opml、docs/SOURCE_COVERAGE.md"))
s.cell(row=3, column=1, value=(
    "本地直连关闭（DISABLED_SOURCES，共 9 项）：aibreakfast, newsnow, iris, "
    "hugging face blog, openai skills, claude code releases, google deepmind, "
    "google ai blog, microsoft ai blog；云端 GitHub Actions 不受影响，"
    "详见 docs/SOURCE_COVERAGE.md"))
s.cell(row=3, column=1).font = Font(name="微软雅黑", size=9, color="C00000")
s.cell(row=3, column=1).alignment = WRAP

r6 = find_row(s, "内置·精选媒体", col=1)
s.cell(row=r6, column=2, value=10)
s.cell(row=r6, column=3, value=(
    "公开 RSS/Atom + 每源条数上限与关键词过滤，保留 30 天窗口；"
    "2026-08-31 新增 OpenRouter、量子位、新智元"))
r9 = find_row(s, "辅助服务·非信源", col=1)
s.cell(row=r9, column=2, value=3)

total_row = find_row(s, "合计", col=1)
s.insert_rows(total_row, amount=2)
new_rows = [
    ("高级信源·默认关闭", 4, "付费/需密钥源：ENABLED 类环境变量 + 密钥开启；默认关闭"),
    ("已评估未接入", 5, "评估后未接入（含 2026-08-31 机器之心；OpenRouter/量子位已转为内置）"),
]
for i, (cat, count, desc) in enumerate(new_rows):
    ridx = total_row + i
    s.cell(row=ridx, column=1, value=cat)
    s.cell(row=ridx, column=2, value=count)
    s.cell(row=ridx, column=3, value=desc)
    for c in range(1, 4):
        cobj = s.cell(row=ridx, column=c)
        cobj.font, cobj.border, cobj.alignment = BODY_FONT, BORDER, WRAP
        cobj.fill = PatternFill("solid", fgColor=COLORS[cat])
s.cell(row=total_row + 2, column=2, value=f"=SUM(B5:B{total_row + 1})")

# ---- 7. 接入方式速查: mention the new curated feeds ----
q = wb["接入方式速查"]
qr = find_row(q, "官方 RSS / Atom", col=1)
old = str(q.cell(row=qr, column=3).value or "")
q.cell(row=qr, column=3, value=f"{old}、量子位、新智元、OpenRouter Announcements")

wb.save(PATH)
print("patched:", PATH, "| detail rows:", ws.max_row - 1)
